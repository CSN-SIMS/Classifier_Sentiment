#Lukas and Jenjira
#Created by following¨and making neacecary changes to a guide at https://pythonprogramming.net
#Work in progress
import os

import nltk
import random
import nltk as nltk
from nltk.corpus import movie_reviews, stopwords
import pickle
from nltk.classify.scikitlearn import SklearnClassifier
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from sklearn.naive_bayes import MultinomialNB,BernoulliNB
from sklearn.linear_model import LogisticRegression,SGDClassifier
from sklearn.svm import SVC, LinearSVC, NuSVC
import os
from nltk.tokenize import word_tokenize
import re
from nltk.classify import ClassifierI
from statistics import mode
from googletrans import Translator
from collections import Counter

#Ärver från ClassifierI från NLTK, används för att räkna vad alla algoritmer bedömmer och låter varje algoritm 'rösta' på resultatet, 7 algoritmer från scikit används i implementeringen.
class VoteClassifier(ClassifierI):
    def __init__(self, *classifiers):
        self._classifiers = classifiers

    def classify(self, features):
        votes = []
        for c in self._classifiers:
            v = c.classify(features)
            votes.append(v)
        # Utkommenterat för att returnera neutralt när confidence är lågt.
        # if(mode(votes) == "neg" and votes.count('pos') == 3):
        # return "Neutral"
        # elif(mode(votes) == "pos" and votes.count("neg") == 3):
        # return "Neutral"

        #Ifall vi bedömmer att algoritmen för ofta väljer neg över pos.
        #if(mode(votes) == "neg" and votes.count('neg') == 4):
            #return 'pos'
        return mode(votes)

    def confidence(self, features):
        votes = []
        for c in self._classifiers:
            v = c.classify(features)
            votes.append(v)
        choice_votes = votes.count(mode(votes))
        conf = choice_votes / len(votes)
        return conf

#Filename syntax example /picklefiles/features.pickle
def loadPickleFile(filename):
    file = open(filename, "rb")
    returnObj = pickle.load(file)
    file.close()
    return returnObj

def savePickle(saveObj, filename):
    save_obj = open(filename, "wb")
    pickle.dump(saveObj, save_obj)
    save_obj.close()

#Print all classifications and their confidence, ändra lidl...
def printVoteConfidence(voted_classifier, testingset):
    lidl = 0
    for i in testingset:
        print("Classification:", voted_classifier.classify(i[0]), "Confidence %:",voted_classifier.confidence(i[0])*100)
        lidl+=1
        if(lidl > 39):
            return

def loadPremadeMovieReviews():
    all_words = []

    for w in movie_reviews.words():
        all_words.append(w.lower())
    all_words = nltk.FreqDist(all_words)
    word_features = list(all_words.keys())[:3000]
    return word_features

def find_features(document, wordfeatures):
    words = word_tokenize(document)
    features = {}
    for w in wordfeatures:
        features[w] = (w in words)
    return features


def loadDatasetFromSingleFiles(posfile, negfile):
    short_pos = open(posfile, "r", encoding = "ISO-8859-1").read()
    short_neg = open(negfile, "r", encoding = "ISO-8859-1").read()

    documents = []
    all_words = []

    allowed_word_types = ["J"]

    for r in short_pos.split('\n'):
        #cleaned = r.replace('\n', ' ')
        cleaned = re.sub('[^a-zA-Z ,.]+', '', r)
        cleaned = cleaned.lower()

        documents.append((cleaned, "pos"))
        words = word_tokenize(cleaned)
        pos = nltk.pos_tag(words)
        for w in pos:
            if w[1][0] in allowed_word_types:
                all_words.append(w[0].lower())

    for r in short_neg.split('\n'):
        # cleaned = r.replace('\n', ' ')
        cleaned = re.sub('[^a-zA-Z ,.]+', '', r)
        cleaned = cleaned.lower()

        documents.append((cleaned, "neg"))
        words = word_tokenize(cleaned)
        pos = nltk.pos_tag(words)
        for w in pos:
            if w[1][0] in allowed_word_types:
                all_words.append(w[0].lower())


    #short_pos_words = word_tokenize(short_pos)
    #short_neg_words = word_tokenize(short_neg)

    all_words = nltk.FreqDist(all_words)
    word_features = list(all_words.keys())[:10000]
    savePickle(word_features, "word_features.pickle")
    savePickle(documents, "documents.pickle")

    featuresets = [(find_features(rev, word_features), category) for (rev,category) in documents]
    random.shuffle(featuresets)
    return featuresets

def sentiment(text, voted_classifier, wordfeatures):
    feats = find_features(text, wordfeatures)
    return voted_classifier.classify(feats), voted_classifier.confidence(feats)

#Används för att analysera en lista av medelanden inladdad av loadTextfilesToList och returnera en lista av filnamnen och [result, confidence]
def analyseListOfMessages(sentimentResults, wordfeatures, voted_classifier):
    sentimentResultList = []

    #Efter att produkten funkar på något sätt, experimentera med dessa för bättre resultat.
    for filename, message in sentimentResults:
        message = message.replace('\n', ' ')
        message = re.sub('[^a-zA-Z ,.]+', '', message)
        message = message.lower()

        result = sentiment(message, voted_classifier, wordfeatures)
        print("Judgement and confidence of: ", filename,  result)
        temp = [filename, result]
        sentimentResultList.append(temp)
    return sentimentResultList

#Testingfunktion för att printa en lista av filnamn och bedömningar
def printSentimentList(sentimentList):
    print("blablabla")
    for result in sentimentList:
        print(result[0], " classified as: ", result[1])

#Testingfunktion för att ladda och spara
def trainAndPickleAllClassifiers(training_set):

    classifier = nltk.NaiveBayesClassifier.train(training_set)
    savePickle(classifier, "picklefiles_eng/basicClassifier.pickle")
    print("Basic classifier saved")

    MNB_classifier = SklearnClassifier(MultinomialNB())
    MNB_classifier.train(training_set)
    savePickle(MNB_classifier, "picklefiles_eng/MNBClassifier.pickle")
    print("MNB classifier saved")

    BernoulliNB_classifier = SklearnClassifier(BernoulliNB())
    BernoulliNB_classifier.train(training_set)
    savePickle(BernoulliNB_classifier, "picklefiles_eng/BNBClassifier.pickle")
    print("BNB classifier saved")

    LogisticRegression_classifier = SklearnClassifier(LogisticRegression(solver='liblinear'))
    LogisticRegression_classifier.train(training_set)
    savePickle(LogisticRegression_classifier, "picklefiles_eng/LRClassifier.pickle")
    print("LRC classifier saved")

    SGDClassifier_classifier = SklearnClassifier(SGDClassifier())
    SGDClassifier_classifier.train(training_set)
    savePickle(SGDClassifier_classifier, "picklefiles_eng/SGDClassifier.pickle")
    print("SGD classifier saved")

    LinearSVC_classifier = SklearnClassifier(LinearSVC(max_iter=10000))
    LinearSVC_classifier.train(training_set)
    savePickle(LinearSVC_classifier, "picklefiles_eng/LinearSVCClassifier.pickle")
    print("LinearSVC classifier saved")

    NuSVC_classifier = SklearnClassifier(NuSVC(gamma='auto'))
    NuSVC_classifier.train(training_set)
    savePickle(NuSVC_classifier, "picklefiles_eng/NUSVCClassifier.pickle")
    print("NuSVC classifier saved")

def loadAndPickleDataset(posfile, negfile):

    featuresets = loadDatasetFromSingleFiles(posfile, negfile)
    trainingSet = featuresets[:10000]
    testingSet = featuresets[10000:]
    savePickle(featuresets, "featuresets.pickle")
    savePickle(trainingSet, "trainingset.pickle")
    savePickle(testingSet, "testingset.pickle")


def pickleAllSwedishClassifiers(basicC, NuSVC, BernouliC, LinearSVC, SGDC, MNBC, LRC):
    savePickle(basicC, "picklefiles_swe/basicClassifier_swe.pickle")
    savePickle(NuSVC, "picklefiles_swe/NuSVC_swe.pickle")
    savePickle(BernouliC, "picklefiles_swe/Bernouli_swe.pickle")
    savePickle(LinearSVC, "picklefiles_swe/LinearSVC_swe.pickle")
    savePickle(SGDC, "picklefiles_swe/SGD_classifier_swe.pickle")
    savePickle(MNBC, "picklefiles_swe/MultinomialNB_swe.pickle")
    savePickle(LRC, "picklefiles_swe/LogisticRegression_swe.pickle")

def translateMessageListToEnglish(messageList):
    translator = Translator()
    translatedMessagelist = []
    filenames = []
    for filename, message in messageList:
        translatedMessagelist.append(translator.translate(message, dest ='en', src='sv').text)
        filenames.append(filename)

    finalList = list(zip(filenames,translatedMessagelist))
    return finalList

def loadTextfilesToList(directory):
    messages = []
    fileNames = []
    files = os.listdir(directory)
    for file in files:
        if(file.endswith(".txt")):
            fileNames.append(file)

    files = [open(directory + "/" + f, 'r').read() for f in files if not f.startswith('.')]
    for p in files:
        messages.append(p)

    finalList = list(zip(fileNames, messages))

    return finalList

def saveExcelFormat(judgementList, category, percentages, filename, append = False):

    if(append):
        wb = load_workbook(filename)
        ws = wb.get_sheet_by_name(category)
        y = ws.max_row
        print(y)
        y += 1
        print(y)
        for fn, judgement in judgementList:
            ws.cell(y, 1, fn)
            ws.cell(y, 2, judgement[0])
            ws.cell(y, 3, judgement[1])
            y += 1
        wb.save(filename)

    else:
        wb = load_workbook(filename)
        ws = wb.create_sheet(category, 0)
        ws.cell(1,1,"Filename")
        ws.cell(1,2, "Judgement")
        ws.cell(1,3, "Confidence %")
        y = 1
        for sent, conf in percentages:
            ws.cell(y,4, percentages[y-1][0] + " " + str(percentages[y-1][1]) + "%")
            y += 1
        y = 2
        for fn, judgement in judgementList:
            ws.cell(y,1, fn)
            ws.cell(y,2, judgement[0])
            ws.cell(y,3, judgement[1] * 100)
            y +=1
        print("Max row:", ws.max_row)
        print("Max column:", ws.max_column)
        wb.save(filename)


def calculatePercentagesOfList(judgementList):
    #negcount, poscount, neucount = 0

    newJudgementList = []
    for filename, judgement in judgementList:
        newJudgementList.append(judgement[0])

    c = Counter(newJudgementList)
    percentages = [(i, c[i] / len(newJudgementList) * 100) for i, count in c.most_common()]
    return percentages
